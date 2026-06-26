"""Excel dashboard rendering."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aios.reports.models import PresentationContext


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="111827")
LABEL_FONT = Font(bold=True, color="374151")
THIN_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))


def write_investment_dashboard(
    context: PresentationContext,
    output_path: str | Path,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    indicators = workbook.create_sheet("Key Indicators")
    reasons = workbook.create_sheet("Reasons")

    _render_dashboard_sheet(dashboard, context)
    _render_indicators_sheet(indicators, context)
    _render_reasons_sheet(reasons, context)

    workbook.save(path)
    return path


def _render_dashboard_sheet(sheet, context: PresentationContext) -> None:
    decision = context.decision
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "AIOS Daily Decision Dashboard"
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Date", context.date),
        ("Data Source", context.metadata.data_source),
        ("Last Update", context.metadata.last_update),
        ("Data Quality", context.metadata.data_quality),
        ("Today's Recommendation", decision.recommendation),
        ("Confidence", decision.confidence),
        ("Risk Level", decision.risk_level.value),
        ("Market Mode", decision.market_mode.value),
        ("Current Position", decision.current_position),
        ("Suggested Position", decision.suggested_position),
        ("Position Delta", decision.position_delta),
        ("Relative Ratio", context.basket.relative_ratio),
        ("Risk Score", context.basket.risk_score),
    ]
    _write_key_value_rows(sheet, rows, start_row=3)

    sheet["D3"] = "Top Reasons"
    sheet["D3"].font = WHITE_FONT
    sheet["D3"].fill = HEADER_FILL
    for index, reason in enumerate(context.top_reasons, start=4):
        sheet.cell(row=index, column=4, value=reason)
        sheet.cell(row=index, column=4).alignment = Alignment(wrap_text=True)

    sheet["D11"] = "Key Indicators"
    sheet["D11"].font = WHITE_FONT
    sheet["D11"].fill = HEADER_FILL
    for index, indicator in enumerate(context.key_indicators[:6], start=12):
        sheet.cell(row=index, column=4, value=indicator.label)
        sheet.cell(row=index, column=5, value=indicator.display_value)

    _style_sheet(sheet)
    _apply_dashboard_conditional_formatting(sheet)
    _add_indicator_chart(sheet, context)
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = "A3:B15"


def _render_indicators_sheet(sheet, context: PresentationContext) -> None:
    sheet.append(["Indicator", "Value", "Display Value"])
    for indicator in context.key_indicators:
        sheet.append([indicator.label, indicator.value, indicator.display_value])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    _style_table_sheet(sheet)
    _add_indicator_color_scale(sheet)


def _render_reasons_sheet(sheet, context: PresentationContext) -> None:
    sheet.append(["Rank", "Reason"])
    for index, reason in enumerate(context.top_reasons, start=1):
        sheet.append([index, reason])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
    _style_table_sheet(sheet)
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True)


def _write_key_value_rows(sheet, rows: list[tuple[str, object]], start_row: int) -> None:
    for offset, (label, value) in enumerate(rows):
        row = start_row + offset
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
        sheet.cell(row=row, column=1).font = LABEL_FONT
        sheet.cell(row=row, column=1).fill = SECTION_FILL
        sheet.cell(row=row, column=1).border = THIN_BORDER
        sheet.cell(row=row, column=2).border = THIN_BORDER


def _style_sheet(sheet) -> None:
    widths = {"A": 24, "B": 22, "C": 4, "D": 34, "E": 18, "F": 16}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 24
    for row in sheet.iter_rows():
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.vertical = "center"
            cell.alignment = alignment


def _style_table_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 22
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def _apply_dashboard_conditional_formatting(sheet) -> None:
    sheet.conditional_formatting.add(
        "B5",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["70"],
            fill=PatternFill("solid", fgColor="DCFCE7"),
        ),
    )
    sheet.conditional_formatting.add(
        "B5",
        CellIsRule(
            operator="lessThan",
            formula=["45"],
            fill=PatternFill("solid", fgColor="FEE2E2"),
        ),
    )


def _add_indicator_color_scale(sheet) -> None:
    if sheet.max_row < 2:
        return
    sheet.conditional_formatting.add(
        f"B2:B{sheet.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="DBEAFE",
            mid_type="percentile",
            mid_value=50,
            mid_color="FEF3C7",
            end_type="max",
            end_color="FEE2E2",
        ),
    )


def _add_indicator_chart(sheet, context: PresentationContext) -> None:
    chart_data_start = 17
    sheet.cell(row=chart_data_start, column=1, value="Chart Metric")
    sheet.cell(row=chart_data_start, column=2, value="Value")
    for index, indicator in enumerate(context.key_indicators[:6], start=chart_data_start + 1):
        sheet.cell(row=index, column=1, value=indicator.label)
        sheet.cell(row=index, column=2, value=indicator.value)

    chart = BarChart()
    chart.title = "Key Indicators"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Metric"
    data = Reference(
        sheet,
        min_col=2,
        min_row=chart_data_start,
        max_row=chart_data_start + min(6, len(context.key_indicators)),
    )
    categories = Reference(
        sheet,
        min_col=1,
        min_row=chart_data_start + 1,
        max_row=chart_data_start + min(6, len(context.key_indicators)),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 12
    sheet.add_chart(chart, "D18")
