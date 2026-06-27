from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from aios.app.models import RunMetadata
from aios.decision.models import (
    BasketSnapshot,
    DecisionResult,
    MarketMode,
    PortfolioPosition,
    RiskLevel,
    TechnicalSnapshot,
)
from aios.reports.models import build_presentation_context, context_to_dict
from aios.reports.presentation import generate_presentation_outputs


def test_context_to_dict_renders_decision_without_recalculating() -> None:
    context = _presentation_context()

    payload = context_to_dict(context)

    assert payload["recommendation"] == "Hold"
    assert payload["confidence"] == 72
    assert payload["risk_level"] == "Low"
    assert payload["market_mode"] == "Uptrend"
    assert payload["current_position"] == 300
    assert payload["suggested_position"] == 300
    assert payload["relative_ratio"] == 1.08
    assert payload["risk_score"] == 12.5
    assert payload["data_source"] == "csv"
    assert payload["last_update"] == "2026-06-26"
    assert payload["data_quality"] == "OK"
    assert payload["manual_mobile_input_used"] is False
    assert payload["latest_manual_input_date"] == "N/A"
    assert payload["manual_tickers_used"] == []
    assert payload["manual_source"] == "None"
    assert payload["data_warnings"] == []
    assert payload["top_reasons"] == [
        "Market classified as Uptrend.",
        "Trend rules favor holding the current position.",
    ]


def test_generate_presentation_outputs(tmp_path: Path) -> None:
    paths = generate_presentation_outputs(
        _presentation_context(),
        tmp_path,
    )

    assert paths.latest_signal.exists()
    assert paths.excel_dashboard.exists()
    assert paths.html_dashboard.exists()

    signal = json.loads(paths.latest_signal.read_text(encoding="utf-8"))
    assert signal["recommendation"] == "Hold"
    assert signal["data_source"] == "csv"
    assert signal["data_quality"] == "OK"
    assert signal["key_indicators"][0]["label"] == "Relative Ratio"

    html = paths.html_dashboard.read_text(encoding="utf-8")
    assert "AIOS Daily Dashboard" in html
    assert "Today's Recommendation" in html
    assert "Hold" in html
    assert "Risk Score" in html
    assert "Data Source" in html
    assert "Data Quality" in html
    assert "Manual Mobile Input" in html
    assert "@media (max-width: 760px)" in html
    assert "<script" not in html.lower()

    workbook = load_workbook(paths.excel_dashboard)
    assert workbook.sheetnames == ["Dashboard", "Key Indicators", "Reasons"]

    dashboard = workbook["Dashboard"]
    assert dashboard["B4"].value == "csv"
    assert dashboard["B5"].value == "No"
    assert dashboard["B6"].value == "None"
    assert dashboard["B7"].value == "N/A"
    assert dashboard["B8"].value == "None"
    assert dashboard["B9"].value == "2026-06-26"
    assert dashboard["B10"].value == "OK"
    assert dashboard["B11"].value == 100
    assert dashboard["B12"].value == 100.0
    assert dashboard["B13"].value is False
    assert dashboard["B14"].value == "Hold"
    assert dashboard["B15"].value == 72
    assert dashboard["B16"].value == "Low"
    assert dashboard["B17"].value == "Uptrend"
    assert dashboard.freeze_panes == "A3"
    assert dashboard.auto_filter.ref == "A3:B24"
    assert len(dashboard._charts) == 1
    assert len(list(dashboard.conditional_formatting)) > 0

    indicators = workbook["Key Indicators"]
    assert indicators.freeze_panes == "A2"
    assert indicators.auto_filter.ref == f"A1:C{indicators.max_row}"
    assert indicators["A2"].value == "Relative Ratio"
    assert indicators["B2"].value == 1.08

    reasons = workbook["Reasons"]
    assert reasons["B2"].value == "Market classified as Uptrend."


def test_data_quality_warnings_are_rendered_for_missing_inputs() -> None:
    context = _presentation_context(
        basket=BasketSnapshot(date="2026-06-26"),
        metadata=RunMetadata(
            data_source="csv",
            provider_used="csv",
            last_update="2026-06-26",
            data_quality="Degraded",
            missing_tickers=["MSFT", "000660.KS"],
            fallback_used=True,
            data_quality_score=40,
            cache_coverage_percentage=50.0,
            recommendation_degraded=True,
        ),
    )

    payload = context_to_dict(context)

    assert payload["risk_score_display"] == "N/A"
    assert payload["relative_ratio_display"] == "N/A"
    assert "MSFT" in payload["data_warnings"][1]
    assert "Risk Score" in payload["data_warnings"][3]


def _presentation_context(
    basket: BasketSnapshot | None = None,
    metadata: RunMetadata | None = None,
):
    decision = DecisionResult(
        date="2026-06-26",
        market_mode=MarketMode.UPTREND,
        recommendation="Hold",
        confidence=72,
        reasons=[
            "Market classified as Uptrend.",
            "Trend rules favor holding the current position.",
        ],
        risk_level=RiskLevel.LOW,
        current_position=300,
        suggested_position=300,
        position_delta=0,
        triggered_rules=["classification.uptrend", "position.hold_trend"],
    )
    basket = basket or BasketSnapshot(
        date="2026-06-26",
        ai_1d=1.2,
        ai_5d=4.1,
        ai_20d=8.8,
        hbm_1d=1.4,
        hbm_5d=5.5,
        hbm_20d=10.2,
        d5=1.4,
        d20=1.4,
        relative_ratio=1.08,
        risk_score=12.5,
    )
    technical = TechnicalSnapshot(
        date="2026-06-26",
        ticker="7709.HK",
        close=120,
        sma_20=110,
        sma_50=100,
        ema_20=112,
        rsi14=62,
        macd=2.4,
        macd_signal=1.2,
        bollinger_upper=130,
        bollinger_lower=95,
        adx14=28,
    )
    portfolio = PortfolioPosition(ticker="7709.HK", current_shares=300)
    return build_presentation_context(
        decision=decision,
        basket=basket,
        technical=technical,
        portfolio=portfolio,
        metadata=metadata or RunMetadata(
            data_source="csv",
            provider_used="csv",
            last_update="2026-06-26",
            data_quality="OK",
            missing_tickers=[],
            data_quality_score=100,
            cache_coverage_percentage=100.0,
        ),
    )
